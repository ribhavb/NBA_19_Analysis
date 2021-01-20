#A brief investigation of team scoring consistency across the pre-bubble 2019-20 NBA Season
#All data used in this program comes from basketball-reference.com
import openpyxl
import statistics
import scipy.stats


nba_teams = ["ATL","BOS","BKN","CHO","CHI","CLE","DAL","DEN",
"DET","GSW","HOU","IND","LAC","LAL","MEM","MIA","MIL","MIN",
"NOP","NYK","OKC","ORL","PHI","PHO","POR","SAC","SAS","TOR",
"UTA","WAS"]


class Team:
    def __init__(self, name, mean_score, dev_score, three_per):
        self.name = name
        self.mean_score = (int (mean_score * 100) / 100)
        self.dev_score = (int (dev_score * 1000) / 1000)
        self.three_per = (int (three_per * 1000) / 1000)

    def __str__(self):
        return f"Team: {self.name}, Mean: {self.mean_score}, Std dev: {self.dev_score}, 3P%: {self.three_per}"




wb = openpyxl.load_workbook("nba19logs.xlsx")
sheets = wb.sheetnames

#initializing lists to populate with our desired data
scores = []
three_attempts = []
three_makes = []

ctr = 0
for ctr in range(30):
    ws = wb[sheets[ctr]]

    scores.append([])
    three_attempts.append(0)
    three_makes.append(0)
    #we want to filter out all games played in the bubble due to uncertainty
    #condition to check if team played in bubble by checking games played 
    if(ws.max_row <= 69):
        for row in range(3, ws.max_row + 1):
            #we use a 2-D array for scores since we eventually want the std. dev. eventually
            scores[ctr].append(int(ws.cell(row,7).value))
            three_attempts[ctr] += (int(ws.cell(row,13).value))
            three_makes[ctr] += (int(ws.cell(row,12).value))
    else:
        for row in range(3, ws.max_row - 7):
            scores[ctr].append(int(ws.cell(row,7).value))
            three_attempts[ctr] += (int(ws.cell(row,13).value))
            three_makes[ctr] += (int(ws.cell(row,12).value))


#initializing lists that hold the final data points we seek to investigate
team_results = []
team_three = []
team_dev = []
ctr = 0
min_dev_team = Team("", 0, 40, 0)
for ctr in range(30):
    name = nba_teams[ctr]
    mean_score = statistics.mean(scores[ctr])
    dev_score = statistics.pstdev(scores[ctr])
    three_per = three_makes[ctr] / three_attempts[ctr]

    team_three.append(three_per)
    team_dev.append(dev_score)

    team = Team(name, mean_score, dev_score, three_per)
    team_results.append(team)

    #quick find of team with lowest std. dev. across the season
    if min_dev_team.dev_score > team.dev_score:
        min_dev_team = team

print(f"The team with the lowest std. dev: {min_dev_team.name}, {min_dev_team.dev_score} \n")

#printing out all our formatted data for each team
for obj in team_results:
    print(obj)

#finding the mean and std. dev. of the deviations themselves
#to see if scoring-deviation across the NBA
team_dev_mean = (int (statistics.mean(team_dev) * 1000) / 1000)
team_dev_dev = (int (statistics.pstdev(team_dev) * 1000) / 1000)

print()
print(f"The average score deviation in the NBA: {team_dev_mean}")
print(f"The deviation of score deviations in the NBA: {team_dev_dev} \n")

#seeing if there exists any correlation between team 3p% and deviation in score
r, p = scipy.stats.pearsonr(team_three, team_dev)

print(f"Pearson correlation: {r}, p value: {p}")
#p-value is too high, so nothing to be concluded here

#next-steps - gather more data across more NBA seasons, and see if the emergence of
#3-point shooting has changed score deviation in any way
